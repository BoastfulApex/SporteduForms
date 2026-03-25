from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.template import loader
from django.db.models import Q
from django.core.paginator import Paginator
from django.shortcuts import redirect, render,get_object_or_404
from apps.authentication.models import *
from apps.main.forms import *
from django.views.generic.edit import CreateView, DeleteView, UpdateView
from django.core.paginator import Paginator
from django.urls import reverse_lazy
from django.contrib import messages
from apps.forms.forms import *
import openpyxl
from django.urls import reverse
from django.db.models import Sum, Max, F



def index(request):
    if not request.user.is_authenticated:
        return redirect('/login/')
    template = 'home/index.html'

    context = {
        'segment': 'dashboard',
    
    }
    html_template = loader.get_template(template)
    return HttpResponse(html_template.render(context, request))


def groups(request):
    if not request.user.is_authenticated:
        return redirect('/login/')
    template = 'home/user/groups/groups.html'
    groups = Group.objects.all()
    paginator = Paginator(groups, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    tashkent_time = timezone.localtime(timezone.now())
    context = {
        'segment': 'groups',
        'page_obj': page_obj,
    }
    html_template = loader.get_template(template)
    return HttpResponse(html_template.render(context, request))


def group_create(request):
    if request.method == "POST":
        print(request.POST)
        form = GroupForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("groups")
        else:
            print(form.errors)
    else:
        form = GroupForm()
    return render(request, "home/user/groups/groups_create.html", {"form": form, 'segment': 'groups'})


def load_months(request):
    year_id = request.GET.get("year_id")
    months = Month.objects.filter(year_id=year_id).order_by("month")
    return JsonResponse(list(months.values("id", "name")), safe=False)

class GroupDelete(DeleteView):
    model = Group
    template_name = 'home/user/groups/group_confirm_delete.html'
    success_url = reverse_lazy('groups')
    

def teacher_create(request):
    if request.method == "POST":
        form = TeacherForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("teachers")
    else:
        form = TeacherForm()
    return render(request, "home/user/teachers/teacher_create.html", {"form": form, 'segment': 'teachers'})


def teachers(request):
    teachers = Teacher.objects.select_related("filial").all()
    template = 'home/user/teachers/teachers.html'
    paginator = Paginator(teachers, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    tashkent_time = timezone.localtime(timezone.now())
    context = {
        'segment': 'teachers',
        'page_obj': page_obj,
    }
    html_template = loader.get_template(template)
    return HttpResponse(html_template.render(context, request))

class TeacherDelete(DeleteView):
    model = Teacher
    template_name = 'home/user/teachers/teacher_confirm_delete.html'
    success_url = reverse_lazy('teachers')
    

def study_module_list(request):

    # Filial bo'yicha modullar
    modules = StudyModule.objects.all()

    # Qidiruv
    search_query = request.GET.get("q", "")
    if search_query:
        modules = modules.filter(
            Q(name__icontains=search_query) |
            Q(study_field__study_field_uz__icontains=search_query) |
            Q(study_field__study_field_ru__icontains=search_query)
        )

    # Paginator
    paginator = Paginator(modules, 50)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Toshkent vaqti
    tashkent_time = timezone.localtime(timezone.now())

    context = {
        "page_obj": page_obj,
        "search_query": search_query,
        "tashkent_time": tashkent_time,
        'segment': 'study_modules',
    }
    return render(request, "home/user/modules/module_list.html", context)


def study_module_create(request):
    if request.method == "POST":
        form = StudyModuleForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("study_module_list")  # Ro'yxatga qaytadi
    else:
        form = StudyModuleForm()

    return render(request, "home/user/modules/module_create.html", {"form": form, 'segment': 'study_modules'})


class StudyModuleDelete(DeleteView):
    model = StudyModule
    template_name = 'home/user/modules/study_module_confirm_delete.html'
    success_url = reverse_lazy('study_module_list')
    

def study_module_upload(request):
    if request.method == "POST":
        form = StudyModuleUploadForm(request.POST, request.FILES)
        if form.is_valid():
            print(request.FILES["file"])
            study_field = form.cleaned_data["study_field"]
            filial = form.cleaned_data["filial"]
            file = request.FILES["file"]

            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active

                count = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):  # 1-qatorda sarlavha
                    name = row[0]  # 1-ustunda modul nomi
                    if name:
                        StudyModule.objects.create(
                            name=name,
                            filial=filial,
                            study_field=study_field
                        )
                        count += 1

                messages.success(request, f"{count} ta modul muvaffaqiyatli qo'shildi!")
                return redirect("study_module_list")

            except Exception as e:
                print(str(e))
                messages.error(request, f"Xato: {str(e)}")

    else:
        form = StudyModuleUploadForm()

    return render(request, "home/user/modules/module_upload.html", {"form": form, 'segment': 'study_modules'})


def group_modules(request, group_id):
    group = get_object_or_404(Group, id=group_id)

    # Guruhga bog'langan modullarni olish
    modules = GroupModuleTeacher.objects.filter(group=group).select_related("teacher", "study_module")

    # Agar hech qanday GroupModuleTeacher yo'q bo'lsa
    if not modules.exists():
        # Guruhning study_field modullarini olish
        study_modules = StudyModule.objects.filter(study_field=group.study_field)

        new_links = []
        for sm in study_modules:
            gmt = GroupModuleTeacher.objects.create(
                teacher=None,  # hozircha teacher tanlanmagan
                study_module=sm,
                group=group
            )
            new_links.append(gmt)

        modules = new_links  # yangi yaratilganlarni ishlatamiz

    return render(request, "home/user/groups/group_modules.html", {
        "group": group,
        "modules": modules,
        'segment': 'groups',
    })
    
    

def assign_teacher(request, group_module_id):
    gmt = get_object_or_404(GroupModuleTeacher, id=group_module_id)

    if request.method == "POST":
        form = AssignTeacherForm(request.POST, instance=gmt)
        if form.is_valid():
            form.save()
            return redirect("group_modules", group_id=gmt.group.id)
    else:
        form = AssignTeacherForm(instance=gmt)

    return render(request, "home/user/groups/assign_teacher.html", {
        "form": form,
        "gmt": gmt,
        'segment': 'groups',
    })
    

def toggle_group_module_teacher_active(request, pk):
    instance = get_object_or_404(GroupModuleTeacher, pk=pk)
    if request.method == "POST":
        form = GroupModuleTeacherActiveForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            return redirect("group_modules", group_id=instance.group.id)
    else:
        form = GroupModuleTeacherActiveForm(instance=instance)
    return render(request, "home/user/groups/edit_active.html", {"form": form, "instance": instance, 'segment': 'groups'})



def form_category_list(request):
    search_query = request.GET.get('q', '')
    form_categories = FormCategory.objects.all()

    if search_query:
        form_categories = form_categories.filter(
            Q(name_uz__icontains=search_query) | Q(name_ru__icontains=search_query)
        )

    paginator = Paginator(form_categories, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'home/user/formcategory/form_category_list.html', {
        'page_obj': page_obj,
        'search_query': search_query,
        'segment': 'form_categories'
    })


def form_category_create(request):
    if request.method == 'POST':
        form = FormCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('form_category_list')
    else:
        form = FormCategoryForm()
    return render(request, 'home/user/formcategory/form_category_form.html', {'form': form, 'title': "Yangi kategoriya qo'shish", 'segment': 'form_categories'})


def form_category_update(request, pk):
    category = get_object_or_404(FormCategory, pk=pk)
    if request.method == 'POST':
        form = FormCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('form_category_list')
    else:
        form = FormCategoryForm(instance=category)
    return render(request, 'home/user/formcategory/form_category_form.html', {'form': form, 'title': 'Kategoriyani tahrirlash', 'segment': 'form_categories'})



def form_category_delete(request, pk):
    category = get_object_or_404(FormCategory, pk=pk)
    if request.method == 'POST':
        category.delete()
        return redirect('form_category_list')
    return render(request, 'home/user/formcategory/form_category_confirm_delete.html', {'category': category, 'segment': 'form_categories'})



def question_list(request):
    search_query = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    active_filter = request.GET.get('active', '')

    questions = Question.objects.select_related('form_category').order_by('-id')

    if search_query:
        questions = questions.filter(
            Q(question_uz__icontains=search_query) |
            Q(question_ru__icontains=search_query)
        )
    if category_id:
        questions = questions.filter(form_category_id=category_id)
    if active_filter == '1':
        questions = questions.filter(active=True)
    elif active_filter == '0':
        questions = questions.filter(active=False)

    paginator = Paginator(questions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'category_id': category_id,
        'active_filter': active_filter,
        'categories': FormCategory.objects.all(),
        'segment': 'questions'
    }
    return render(request, "home/user/question/question_list.html", context)

def question_sample_download(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Savollar"
    headers = [
        "Savol (UZ)", "Savol (RU)",
        "Javob 1 (UZ)", "Javob 1 (RU)", "Ball 1",
        "Javob 2 (UZ)", "Javob 2 (RU)", "Ball 2",
        "Javob 3 (UZ)", "Javob 3 (RU)", "Ball 3",
        "Javob 4 (UZ)", "Javob 4 (RU)", "Ball 4",
    ]
    ws.append(headers)
    ws.append([
        "O'qituvchi darsni qanday tushuntiradi?",
        "Как учитель объясняет урок?",
        "Juda yaxshi", "Очень хорошо", 5,
        "Yaxshi", "Хорошо", 4,
        "Qoniqarli", "Удовлетворительно", 3,
        "Yomon", "Плохо", 1,
    ])
    ws.append([
        "Dars qiziqarlimi?",
        "Интересен ли урок?",
        "Ha, juda qiziqarli", "Да, очень интересно", 5,
        "Qisman", "Частично", 3,
        "Yo'q", "Нет", 1,
        "", "", 0,
    ])
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="savollar_namuna.xlsx"'
    wb.save(response)
    return response


def question_upload(request):
    if request.method == "POST":
        form = QuestionUploadForm(request.POST, request.FILES)
        if form.is_valid():
            form_category = form.cleaned_data["form_category"]
            file = request.FILES["file"]
            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                count = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Ustunlar: question_uz, question_ru, a1_uz, a1_ru, score1, a2_uz, a2_ru, score2, a3_uz, a3_ru, score3, a4_uz, a4_ru, score4
                    if not row[0]:
                        continue
                    question = Question.objects.create(
                        form_category=form_category,
                        question_uz=row[0] or "",
                        question_ru=row[1] or "",
                        active=True,
                    )
                    # Har bir javob (4 ta): (uz, ru, ball) -> 3 ta ustun
                    for i in range(4):
                        base = 2 + i * 3
                        ans_uz = row[base] if len(row) > base else None
                        ans_ru = row[base + 1] if len(row) > base + 1 else None
                        score = row[base + 2] if len(row) > base + 2 else 0
                        if ans_uz:
                            Answer.objects.create(
                                question=question,
                                answer_uz=ans_uz,
                                answer_ru=ans_ru or ans_uz,
                                score=score or 0,
                            )
                    count += 1
                messages.success(request, f"{count} ta savol muvaffaqiyatli qo'shildi!")
                return redirect("question_list")
            except Exception as e:
                messages.error(request, f"Xato: {str(e)}")
    else:
        form = QuestionUploadForm()
    return render(request, "home/user/question/question_upload.html", {"form": form, "segment": "questions"})


def question_create(request):
    if request.method == "POST":
        form = QuestionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("question_list")
    else:
        form = QuestionForm()
    return render(request, "home/user/question/question_create.html", {"form": form, 'segment': 'questions'})


def question_update(request, pk):
    question = get_object_or_404(Question, pk=pk)
    if request.method == "POST":
        form = QuestionForm(request.POST, instance=question)
        if form.is_valid():
            form.save()
            return redirect("question_list")
    else:
        form = QuestionForm(instance=question)
    return render(request, "home/user/question/question_edit.html", {"form": form, "question": question, 'segment': 'questions'})


def question_delete(request, pk):
    question = get_object_or_404(Question, pk=pk)
    if request.method == "POST":
        question.delete()
        return redirect("question_list")
    return render(request, "home/user/question/question_delete.html", {"object": question, 'segment': 'questions'})


def answer_list(request):
    search_query = request.GET.get('q', '')
    answers = Answer.objects.all()

    if search_query:
        answers = answers.filter(
            Q(answer_uz__icontains=search_query) |
            Q(question__question_uz__icontains=search_query)
        )

    paginator = Paginator(answers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "home/user/answers/answers_list.html", {
        "page_obj": page_obj,
        "search_query": search_query,
        "segment": "answers"
    })


def answer_create(request):
    if request.method == "POST":
        form = AnswerForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("answer_list")
    else:
        form = AnswerForm()
    return render(request, "home/user/answers/answer_form.html", {"form": form, "segment": "answers", "title": "Yangi javob qo'shish"})


def answer_edit(request, pk):
    answer = get_object_or_404(Answer, pk=pk)
    if request.method == "POST":
        form = AnswerForm(request.POST, instance=answer)
        if form.is_valid():
            form.save()
            return redirect("answer_list")
    else:
        form = AnswerForm(instance=answer)
    return render(request, "home/user/answers/answer_form.html", {"form": form,"segment": "answers", "title": "Javobni tahrirlash"})


def answer_delete(request, pk):
    answer = get_object_or_404(Answer, pk=pk)
    if request.method == "POST":
        answer.delete()
        return redirect("answer_list")
    return render(request, "home/user/answers/answer_delete.html", {"answer": answer, "segment": "answers"})


# def add_answer_view(request):
#     """
#     Bitta question tanlanadi va bir nechta javoblar formset orqali kiritiladi.
#     """
#     title = "Javoblar qo'shish"
#     question_form = QuestionSelectForm(request.POST or None)
#     formset = AddAnswerFormSet(request.POST or None, queryset=Answer.objects.none())

#     if request.method == "POST":
#         if question_form.is_valid() and formset.is_valid():
#             question = question_form.cleaned_data['question']

#             for form in formset:
#                 if form.cleaned_data and not form.cleaned_data.get('DELETE'):
#                     answer = form.save(commit=False)
#                     answer.question = question
#                     answer.save()

#             return redirect('answer_list')
#         else:
#             print(question_form.errors, formset.errors)
#     context = {
#         "title": title,
#         "form": question_form,
#         "formset": formset,
#     }
#     return render(request, 'home/user/answers/answer_dynamic_form.html', context)


def add_answer_view(request):
    title = "Javoblar qo'shish"
    question_form = QuestionSelectForm(request.POST or None)
    formset = AddAnswerFormSet(request.POST or None, queryset=Answer.objects.none())

    if request.method == "POST":
        if question_form.is_valid() and formset.is_valid():
            question = question_form.cleaned_data['question']
            for form in formset:
                if form.cleaned_data and not form.cleaned_data.get('DELETE'):
                    answer = form.save(commit=False)
                    answer.question = question
                    answer.save()

            # 🔁 AJAX yuborilgan bo'lsa, JSON qaytaramiz
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    "success": True,
                    "redirect_url": reverse("answer_list")
                })
            else:
                return redirect('answer_list')

        # Agar xato bo'lsa
        elif request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                "success": False,
                "errors": {
                    "question_errors": question_form.errors,
                    "formset_errors": formset.errors
                }
            }, status=400)

    context = {
        "title": title,
        "form": question_form,
        "formset": formset,
        "segment": "answers"
    }
    return render(request, "home/user/answers/answer_dynamic_form.html", context)


def schedule_create(request):
    if request.method == 'POST':
        form = ScheduleForm(request.POST)
        if form.is_valid():
            year = form.cleaned_data['year']
            month = form.cleaned_data['month']
            filial = form.cleaned_data['filial']
            selected_month = form.cleaned_data['month']

            teachers = Teacher.objects.filter(
                teacher_modules__group__month_id=month
            ).distinct()

            report_data = []

            for teacher in teachers:
                teacher_modules = GroupModuleTeacher.objects.filter(teacher=teacher)
                teacher_total_score = 0
                teacher_max_score = 0

                for module in teacher_modules:
                    user_answers = UserAnswer.objects.filter(module=module).all()
                    for ua in user_answers:
                        for ans in ua.answer.all():
                            teacher_total_score += ans.score
                            teacher_max_score += Answer.objects.filter(question=ua.question).aggregate(max_score=Max('score'))['max_score'] or 0
                percent = round((teacher_total_score / teacher_max_score * 100), 2) if teacher_max_score else 0
                report_data.append({
                    'teacher': teacher.name,
                    'total_score': teacher_total_score,
                    'max_score': teacher_max_score,
                    'percent': percent,
                })
            report_data.sort(key=lambda x: x['percent'], reverse=True)
        return render(request, 'home/reports/teacher_report.html', {
            'form': form,
            'report_data': report_data,
            'segment': 'reports'
        })
    else:
        form = ScheduleForm()
    return render(request, 'home/reports/select_date.html', {'form': form, 'segment': 'reports'})


# ===================== YANGI HISOBOT TIZIMI =====================

def report_select(request):
    """So'rovnoma turini tanlash sahifasi"""
    categories = FormCategory.objects.all()
    return render(request, 'home/reports/report_select.html', {
        'categories': categories,
        'segment': 'reports'
    })


def report_detail(request, category_id):
    """Tanlangan so'rovnoma bo'yicha hisobot"""
    from apps.main.models import StudyField
    category = get_object_or_404(FormCategory, pk=category_id)
    questions = Question.objects.filter(form_category=category).prefetch_related('answer_set')

    if category.for_rating:
        # ===== REYTING HISOBOT =====
        from apps.main.forms import MonthSelectForm
        form = MonthSelectForm(request.GET or None)
        report_data = []
        selected_month = None

        if form.is_valid():
            selected_month = form.cleaned_data['month']
            teachers = Teacher.objects.filter(
                teacher_modules__group__month=selected_month
            ).distinct()

            for teacher in teachers:
                teacher_modules = GroupModuleTeacher.objects.filter(
                    teacher=teacher,
                    group__month=selected_month
                )
                total_score = 0
                max_score = 0
                for module in teacher_modules:
                    user_answers = UserAnswer.objects.filter(
                        module=module,
                        question__form_category=category
                    )
                    for ua in user_answers:
                        for ans in ua.answer.all():
                            total_score += ans.score
                        max_score += Answer.objects.filter(
                            question=ua.question
                        ).aggregate(m=Max('score'))['m'] or 0
                if max_score == 0:
                    continue
                percent = round(total_score / max_score * 100, 2)
                report_data.append({
                    'teacher': teacher.name,
                    'total_score': total_score,
                    'max_score': max_score,
                    'percent': percent,
                })
            report_data.sort(key=lambda x: x['percent'], reverse=True)

        return render(request, 'home/reports/report_rating.html', {
            'category': category,
            'form': form,
            'report_data': report_data,
            'selected_month': selected_month,
            'segment': 'reports',
        })

    else:
        # ===== YO'NALISH BO'YICHA HISOBOT =====
        study_fields = StudyField.objects.all()
        report_data = []

        for question in questions:
            answers = Answer.objects.filter(question=question)
            # Barcha UserAnswer'lar bu savol uchun
            total_respondents = UserAnswer.objects.filter(question=question).count()

            answers_data = []
            for answer in answers:
                count = UserAnswer.objects.filter(
                    question=question,
                    answer=answer
                ).count()
                percent = round(count / total_respondents * 100, 1) if total_respondents else 0
                answers_data.append({
                    'answer_uz': answer.answer_uz,
                    'answer_ru': answer.answer_ru,
                    'score': answer.score,
                    'count': count,
                    'percent': percent,
                })

            # Yo'nalishlar bo'yicha breakdown
            fields_data = []
            for sf in study_fields:
                sf_count = UserAnswer.objects.filter(
                    question=question,
                    user__group__study_field=sf
                ).count()
                if sf_count == 0:
                    continue
                sf_answers = []
                for answer in answers:
                    a_count = UserAnswer.objects.filter(
                        question=question,
                        answer=answer,
                        user__group__study_field=sf
                    ).count()
                    a_pct = round(a_count / sf_count * 100, 1) if sf_count else 0
                    sf_answers.append({
                        'answer_uz': answer.answer_uz,
                        'count': a_count,
                        'percent': a_pct,
                    })
                fields_data.append({
                    'study_field': sf.study_field_uz,
                    'total': sf_count,
                    'answers': sf_answers,
                })

            report_data.append({
                'question_uz': question.question_uz,
                'question_ru': question.question_ru,
                'total_respondents': total_respondents,
                'answers': answers_data,
                'fields_data': fields_data,
            })

        return render(request, 'home/reports/report_survey.html', {
            'category': category,
            'report_data': report_data,
            'segment': 'reports',
        })


def report_export_excel(request, category_id):
    from apps.main.models import StudyField
    from apps.main.forms import MonthSelectForm
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    category = get_object_or_404(FormCategory, pk=category_id)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    sub_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(cell):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    def style_cell(cell, align=None):
        cell.alignment = align or left
        cell.border = border

    wb = openpyxl.Workbook()

    if category.for_rating:
        # ===== REYTING EXCEL =====
        form = MonthSelectForm(request.GET or None)
        selected_month = None
        report_data = []

        if form.is_valid():
            selected_month = form.cleaned_data['month']
            teachers = Teacher.objects.filter(
                teacher_modules__group__month=selected_month
            ).distinct()
            for teacher in teachers:
                teacher_modules = GroupModuleTeacher.objects.filter(
                    teacher=teacher, group__month=selected_month
                )
                total_score = 0
                max_score = 0
                for module in teacher_modules:
                    user_answers = UserAnswer.objects.filter(
                        module=module, question__form_category=category
                    )
                    for ua in user_answers:
                        for ans in ua.answer.all():
                            total_score += ans.score
                        max_score += Answer.objects.filter(
                            question=ua.question
                        ).aggregate(m=Max('score'))['m'] or 0
                if max_score == 0:
                    continue
                percent = round(total_score / max_score * 100, 2)
                report_data.append({
                    'teacher': teacher.name,
                    'total_score': total_score,
                    'max_score': max_score,
                    'percent': percent,
                })
            report_data.sort(key=lambda x: x['percent'], reverse=True)

        ws = wb.active
        ws.title = 'Reyting'
        headers = ['#', "O'qituvchi", "To'plagan ball", "Maksimal ball", "Natija (%)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            style_header(cell)
        for i, item in enumerate(report_data, 1):
            row = [i, item['teacher'], item['total_score'], item['max_score'], item['percent']]
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=i + 1, column=col, value=val)
                style_cell(cell, center if col != 2 else left)
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        for c in ['C', 'D', 'E']:
            ws.column_dimensions[c].width = 18
        filename = f"reyting_{category.name_uz}_{selected_month or 'barchasi'}.xlsx"

    else:
        # ===== SO'ROVNOMA EXCEL =====
        study_fields = StudyField.objects.all()
        questions = Question.objects.filter(form_category=category)

        ws = wb.active
        ws.title = 'Hisobot'
        row_num = 1

        for question in questions:
            answers = Answer.objects.filter(question=question)
            total_respondents = UserAnswer.objects.filter(question=question).count()

            # Savol sarlavhasi
            cell = ws.cell(row=row_num, column=1, value=f"Savol: {question.question_uz}")
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.alignment = left
            cell.border = border
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
            row_num += 1

            # Umumiy natija sarlavha
            for col, h in enumerate(['Javob', 'Ball', 'Soni', 'Foiz (%)'], 1):
                cell = ws.cell(row=row_num, column=col, value=h)
                style_header(cell)
            row_num += 1

            for answer in answers:
                count = UserAnswer.objects.filter(question=question, answer=answer).count()
                percent = round(count / total_respondents * 100, 1) if total_respondents else 0
                row_vals = [answer.answer_uz, answer.score, count, percent]
                for col, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=row_num, column=col, value=val)
                    style_cell(cell, center if col != 1 else left)
                row_num += 1

            # Yo'nalishlar bo'yicha
            row_num += 1
            cell = ws.cell(row=row_num, column=1, value="Yo'nalishlar bo'yicha:")
            cell.font = Font(bold=True)
            cell.fill = sub_fill
            cell.border = border
            ans_list = list(answers)
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2 + len(ans_list))
            row_num += 1

            # Yo'nalish sarlavha
            header_row = ["Yo'nalish", "Jami"] + [a.answer_uz[:25] for a in ans_list]
            for col, h in enumerate(header_row, 1):
                cell = ws.cell(row=row_num, column=col, value=h)
                style_header(cell)
            row_num += 1

            for sf in study_fields:
                sf_count = UserAnswer.objects.filter(
                    question=question, user__group__study_field=sf
                ).count()
                if sf_count == 0:
                    continue
                row_vals = [sf.study_field_uz, sf_count]
                for answer in ans_list:
                    a_count = UserAnswer.objects.filter(
                        question=question, answer=answer, user__group__study_field=sf
                    ).count()
                    a_pct = round(a_count / sf_count * 100, 1) if sf_count else 0
                    row_vals.append(f"{a_count} ({a_pct}%)")
                for col, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=row_num, column=col, value=val)
                    style_cell(cell, center if col != 1 else left)
                row_num += 1

            row_num += 2  # Bo'sh qator savollar orasida

        ws.column_dimensions['A'].width = 40
        for i in range(2, 10):
            ws.column_dimensions[get_column_letter(i)].width = 20
        filename = f"sorovnoma_{category.name_uz}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def ajax_load_months(request):
    year_id = request.GET.get('year_id')
    months = Month.objects.filter(year_id=year_id).values('id', 'name')
    return JsonResponse(list(months), safe=False)